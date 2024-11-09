from openpyxl import load_workbook

def file_data_extract(file_path):
    # Load the workbook
    workbook = load_workbook(file_path)

    # Define the list of keywords to check for
    keywords = ['reg no', 'reg', 'roll no', 'roll', 'registration']

    # Initialize a list to hold the extracted data from all sheets
    all_column_data = []

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate through the header row to find the column containing any of the keywords
        reg_column_index = None
        for col_index, col in enumerate(worksheet.iter_cols(min_row=1, max_row=1, values_only=True), start=1):
            for cell in col:
                if cell and any(keyword in str(cell).lower() for keyword in keywords):
                    reg_column_index = col_index
                    break
            if reg_column_index:
                break

        # If the column is found, extract the data from this sheet
        if reg_column_index:
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                           min_col=reg_column_index, max_col=reg_column_index,
                                           values_only=True):
                value = row[0]
                if value is not None:  # Only add non-empty cells
                    # Strip whitespace before appending
                    all_column_data.append(str(value).strip())

    # If data is collected, join it into a comma-separated string
    if all_column_data:
        return ', '.join(all_column_data)
    else:
        return "Column with specified keywords not found."

# Test the file_data_extract function
# file_path = "my_sample_data/multi_page.xlsx"  # Replace with your actual file path
#
# try:
#     result = file_data_extract(file_path)
#     print("Extracted Data:", result)
# except Exception as e:
#     print("An error occurred:", str(e))
#
# fetch_and_update_users("Meow tech", result)
